from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from .models import Product
from .forms import ProductForm
import pandas as pd
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
import datetime

@login_required
def product_list(request):
    products = Product.objects.all()
    return render(request, 'products/product_list.html', {'products': products})

@login_required
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('product_list')
    else:
        form = ProductForm()
    return render(request, 'products/add_product.html', {'form': form})

@login_required
def import_products(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active
        rows = ws.iter_rows(min_row=2, values_only=True)
        products_created = 0
        products_updated = 0
        for row in rows:
            name, description, price, quantity = row
            product, created = Product.objects.update_or_create(
                name=name,
                defaults={
                    'description': description,
                    'price': price,
                    'quantity': quantity
                }
            )
            if created:
                products_created += 1
            else:
                products_updated += 1
        messages.success(request, f'Imported {products_created} new products and updated {products_updated} products.')
        return redirect('products:list')
    return render(request, 'products/import.html')

def export_products_to_excel(request):
    products = Product.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = ['Name', 'Description', 'Price', 'Quantity']
    ws.append(headers)
    for product in products:
        ws.append([product.name, product.description, product.price, product.quantity])
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=products_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response

class ProductListView(ListView):
    model = Product
    template_name = 'products/index.html'
    context_object_name = 'products'

class ProductDetailView(DetailView):
    model = Product
    template_name = 'products/detail.html'

class ProductCreateView(CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'products/create.html'
    success_url = reverse_lazy('products:list')

    def form_valid(self, form):
        # يمكنك إضافة أي منطق إضافي هنا إذا لزم الأمر
        return super().form_valid(form)

class ProductUpdateView(UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'products/update.html'
    success_url = reverse_lazy('products:list')